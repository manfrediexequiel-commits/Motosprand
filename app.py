import streamlit as st
import pandas as pd
import sqlite3
import hashlib
import secrets
import re
import os
import json
import smtplib
import requests
import shutil
from datetime import datetime, timedelta
from difflib import get_close_matches
from fpdf import FPDF
from contextlib import contextmanager
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple, Any
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
import threading
import time
import schedule
import mercadopago
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import qrcode
from io import BytesIO
import base64

# =============================================================================
# CONFIGURACIÓN INICIAL Y VARIABLES DE ENTORNO
# =============================================================================

st.set_page_config(
    page_title="SAT Pro Enterprise - Gestión Profesional de Taller",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://satpro.com/soporte',
        'Report a bug': 'https://satpro.com/bug',
        'About': "SAT Pro Enterprise v3.0 - Sistema de Gestión Integral para Talleres Mecánicos"
    }
)

# Cargar configuración desde variables de entorno o secrets
class Config:
    """Gestión centralizada de configuración"""
    # Base de datos
    DB_PATH = os.getenv('DB_PATH', 'gestion_taller_enterprise.db')
    BACKUP_DIR = os.getenv('BACKUP_DIR', './backups')
    
    # Email (SMTP)
    SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
    SMTP_USER = os.getenv('SMTP_USER', '')
    SMTP_PASS = os.getenv('SMTP_PASS', '')
    EMAIL_FROM = os.getenv('EMAIL_FROM', 'sistema@satpro.com')
    
    # WhatsApp Business API (Twilio o similar)
    WHATSAPP_API_KEY = os.getenv('WHATSAPP_API_KEY', '')
    WHATSAPP_API_URL = os.getenv('WHATSAPP_API_URL', '')
    
    # MercadoPago
    MP_ACCESS_TOKEN = os.getenv('MP_ACCESS_TOKEN', '')
    MP_PUBLIC_KEY = os.getenv('MP_PUBLIC_KEY', '')
    
    # Cloud/Deploy
    CLOUD_PROVIDER = os.getenv('CLOUD_PROVIDER', 'local')  # local, aws, gcp, azure
    DOCKER_MODE = os.getenv('DOCKER_MODE', 'false').lower() == 'true'
    
    # Seguridad
    SECRET_KEY = os.getenv('SECRET_KEY', secrets.token_hex(32))
    SESSION_TIMEOUT = int(os.getenv('SESSION_TIMEOUT', '480'))  # 8 horas en minutos

config = Config()

# Crear directorios necesarios
Path(config.BACKUP_DIR).mkdir(exist_ok=True)
Path('./temp').mkdir(exist_ok=True)
Path('./exports').mkdir(exist_ok=True)

# CSS Personalizado avanzado
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * { font-family: 'Inter', sans-serif; }
    
    .main-header { 
        color: #1a73e8; 
        font-size: 2.5rem; 
        font-weight: 700;
        letter-spacing: -0.5px;
    }
    
    .metric-card { 
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px; 
        border-radius: 16px; 
        color: white;
        box-shadow: 0 10px 25px rgba(0,0,0,0.1);
        transition: transform 0.2s;
    }
    .metric-card:hover { transform: translateY(-5px); }
    
    .status-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .status-entregado { background: #d4edda; color: #155724; }
    .status-reparacion { background: #fff3cd; color: #856404; }
    .status-urgente { background: #f8d7da; color: #721c24; }
    .status-espera { background: #d1ecf1; color: #0c5460; }
    
    .stButton>button {
        border-radius: 8px;
        font-weight: 500;
        transition: all 0.2s;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# SISTEMA DE BASE DE DATOS ENTERPRISE
# =============================================================================

@dataclass
class DatabaseManager:
    """Gestor de base de datos con pooling y replicación"""
    
    def __post_init__(self):
        self._local = threading.local()
        self.init_database()
        self.start_backup_scheduler()
    
    @property
    def connection(self):
        """Thread-safe connection pooling"""
        if not hasattr(self._local, 'conn') or self._local.conn is None:
            self._local.conn = sqlite3.connect(
                config.DB_PATH,
                check_same_thread=False,
                timeout=30,
                isolation_level=None
            )
            self._local.conn.execute("PRAGMA foreign_keys = ON;")
            self._local.conn.execute("PRAGMA journal_mode = WAL;")
            self._local.conn.execute("PRAGMA synchronous = NORMAL;")
            self._local.conn.row_factory = sqlite3.Row
        return self._local.conn
    
    @contextmanager
    def get_cursor(self):
        """Context manager para transacciones seguras"""
        cursor = self.connection.cursor()
        try:
            yield cursor
            self.connection.commit()
        except Exception as e:
            self.connection.rollback()
            raise e
    
    def init_database(self):
        """Inicialización completa del esquema"""
        with self.get_cursor() as cursor:
            cursor.executescript("""
                -- Inventario con tracking completo
                CREATE TABLE IF NOT EXISTS inventario (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo TEXT UNIQUE NOT NULL,
                    item TEXT NOT NULL,
                    descripcion TEXT,
                    categoria TEXT,
                    cantidad INTEGER NOT NULL DEFAULT 0 CHECK(cantidad >= 0),
                    minimo INTEGER NOT NULL DEFAULT 1,
                    precio_costo REAL DEFAULT 0,
                    precio_venta REAL NOT NULL DEFAULT 0,
                    margen REAL GENERATED ALWAYS AS ((precio_venta - precio_costo) / NULLIF(precio_costo, 0)) STORED,
                    proveedor_id INTEGER,
                    ubicacion TEXT,
                    codigo_barras TEXT,
                    activo BOOLEAN DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(proveedor_id) REFERENCES proveedores(id)
                );
                
                -- Proveedores
                CREATE TABLE IF NOT EXISTS proveedores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL,
                    contacto TEXT,
                    telefono TEXT,
                    email TEXT,
                    direccion TEXT,
                    cuit TEXT UNIQUE,
                    condicion_pago TEXT DEFAULT 'Contado',
                    descuento_default REAL DEFAULT 0,
                    activo BOOLEAN DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                
                -- Órdenes de reparación enterprise
                CREATE TABLE IF NOT EXISTS reparaciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_orden TEXT UNIQUE NOT NULL,
                    fecha_ingreso TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    fecha_estimada DATE,
                    fecha_finalizacion TIMESTAMP,
                    cliente_id INTEGER NOT NULL,
                    vehiculo_id INTEGER,
                    falla TEXT NOT NULL,
                    diagnostico TEXT,
                    solucion_aplicada TEXT,
                    estado TEXT DEFAULT 'Recibido' CHECK(estado IN 
                        ('Recibido', 'Diagnóstico', 'Presupuestado', 'Aprobado', 
                         'Esperando repuestos', 'En reparación', 'Pruebas QC', 
                         'Listo para entrega', 'Entregado', 'Garantía', 'Cancelado')),
                    prioridad TEXT DEFAULT 'Normal',
                    mano_obra REAL DEFAULT 0,
                    descuento_porcentaje REAL DEFAULT 0,
                    descuento_monto REAL DEFAULT 0,
                    subtotal REAL DEFAULT 0,
                    impuestos REAL DEFAULT 0,
                    total REAL DEFAULT 0,
                    metodo_pago TEXT,
                    pagado BOOLEAN DEFAULT 0,
                    tecnico_id INTEGER,
                    notas_internas TEXT,
                    notas_cliente TEXT,
                    kilometraje INTEGER,
                    combustible INTEGER,
                    activo BOOLEAN DEFAULT 1,
                    FOREIGN KEY(cliente_id) REFERENCES clientes(id),
                    FOREIGN KEY(vehiculo_id) REFERENCES vehiculos(id),
                    FOREIGN KEY(tecnico_id) REFERENCES usuarios(id)
                );
                
                -- Clientes con historial
                CREATE TABLE IF NOT EXISTS clientes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL,
                    apellido TEXT,
                    telefono TEXT,
                    email TEXT UNIQUE,
                    direccion TEXT,
                    dni_cuit TEXT UNIQUE,
                    fecha_nacimiento DATE,
                    observaciones TEXT,
                    fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    total_ordenes INTEGER DEFAULT 0,
                    total_gastado REAL DEFAULT 0,
                    ultima_visita DATE,
                    activo BOOLEAN DEFAULT 1
                );
                
                -- Vehículos vinculados a clientes
                CREATE TABLE IF NOT EXISTS vehiculos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    cliente_id INTEGER NOT NULL,
                    patente TEXT NOT NULL,
                    marca TEXT,
                    modelo TEXT,
                    anio INTEGER,
                    color TEXT,
                    numero_chasis TEXT,
                    numero_motor TEXT,
                    kilometraje_actual INTEGER,
                    ultimo_service DATE,
                    observaciones TEXT,
                    activo BOOLEAN DEFAULT 1,
                    FOREIGN KEY(cliente_id) REFERENCES clientes(id),
                    UNIQUE(cliente_id, patente)
                );
                
                -- Historial de consumo con costos
                CREATE TABLE IF NOT EXISTS historial_consumo (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    id_reparacion INTEGER NOT NULL,
                    id_repuesto INTEGER NOT NULL,
                    cantidad INTEGER NOT NULL DEFAULT 1,
                    precio_costo REAL,
                    precio_venta REAL NOT NULL,
                    margen REAL,
                    usuario_id INTEGER,
                    fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(id_reparacion) REFERENCES reparaciones(id) ON DELETE CASCADE,
                    FOREIGN KEY(id_repuesto) REFERENCES inventario(id),
                    FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
                );
                
                -- Historial de estados (auditoría completa)
                CREATE TABLE IF NOT EXISTS historial_estados (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    id_reparacion INTEGER NOT NULL,
                    estado_anterior TEXT,
                    estado_nuevo TEXT NOT NULL,
                    usuario_id INTEGER NOT NULL,
                    fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    nota_cambio TEXT,
                    ip_address TEXT,
                    user_agent TEXT,
                    FOREIGN KEY(id_reparacion) REFERENCES reparaciones(id) ON DELETE CASCADE,
                    FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
                );
                
                -- Usuarios del sistema (seguridad reforzada)
                CREATE TABLE IF NOT EXISTS usuarios (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    salt TEXT NOT NULL,
                    nombre_completo TEXT,
                    email TEXT UNIQUE,
                    rol TEXT DEFAULT 'tecnico' CHECK(rol IN ('admin', 'supervisor', 'tecnico', 'recepcion')),
                    activo BOOLEAN DEFAULT 1,
                    ultimo_acceso TIMESTAMP,
                    intentos_fallidos INTEGER DEFAULT 0,
                    bloqueado_hasta TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                
                -- Notificaciones enviadas
                CREATE TABLE IF NOT EXISTS notificaciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tipo TEXT NOT NULL CHECK(tipo IN ('email', 'whatsapp', 'sms', 'push')),
                    destinatario TEXT NOT NULL,
                    asunto TEXT,
                    contenido TEXT,
                    estado TEXT DEFAULT 'pendiente' CHECK(estado IN ('pendiente', 'enviado', 'error', 'leido')),
                    orden_id INTEGER,
                    error_msg TEXT,
                    enviado_at TIMESTAMP,
                    creado_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(orden_id) REFERENCES reparaciones(id)
                );
                
                -- Órdenes de compra a proveedores
                CREATE TABLE IF NOT EXISTS ordenes_compra (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    proveedor_id INTEGER NOT NULL,
                    numero_oc TEXT UNIQUE NOT NULL,
                    fecha_emision TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    fecha_entrega_esperada DATE,
                    estado TEXT DEFAULT 'Pendiente' CHECK(estado IN ('Pendiente', 'Parcial', 'Completa', 'Cancelada')),
                    subtotal REAL DEFAULT 0,
                    impuestos REAL DEFAULT 0,
                    total REAL DEFAULT 0,
                    notas TEXT,
                    FOREIGN KEY(proveedor_id) REFERENCES proveedores(id)
                );
                
                CREATE TABLE IF NOT EXISTS ordenes_compra_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    orden_compra_id INTEGER NOT NULL,
                    repuesto_id INTEGER NOT NULL,
                    cantidad_solicitada INTEGER NOT NULL,
                    cantidad_recibida INTEGER DEFAULT 0,
                    precio_unitario REAL,
                    FOREIGN KEY(orden_compra_id) REFERENCES ordenes_compra(id) ON DELETE CASCADE,
                    FOREIGN KEY(repuesto_id) REFERENCES inventario(id)
                );
                
                -- Pagos y facturación
                CREATE TABLE IF NOT EXISTS pagos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    orden_id INTEGER NOT NULL,
                    monto REAL NOT NULL,
                    metodo TEXT NOT NULL CHECK(metodo IN ('efectivo', 'tarjeta', 'transferencia', 'mercadopago', 'cheque', 'cuenta_corriente')),
                    referencia TEXT,
                    estado TEXT DEFAULT 'completado' CHECK(estado IN ('pendiente', 'completado', 'rechazado', 'reembolsado')),
                    mercadopago_id TEXT,
                    fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(orden_id) REFERENCES reparaciones(id)
                );
                
                -- Índices de performance
                CREATE INDEX IF NOT EXISTS idx_rep_estado ON reparaciones(estado);
                CREATE INDEX IF NOT EXISTS idx_rep_fecha ON reparaciones(fecha_ingreso);
                CREATE INDEX IF NOT EXISTS idx_rep_cliente ON reparaciones(cliente_id);
                CREATE INDEX IF NOT EXISTS idx_inv_codigo ON inventario(codigo);
                CREATE INDEX IF NOT EXISTS idx_inv_categoria ON inventario(categoria);
                CREATE INDEX IF NOT EXISTS idx_hist_fecha ON historial_consumo(fecha);
                CREATE INDEX IF NOT EXISTS idx_notif_estado ON notificaciones(estado);
            """)
            
            # Insertar usuario admin por defecto si no existe
            cursor.execute("SELECT COUNT(*) FROM usuarios WHERE username = 'admin'")
            if cursor.fetchone()[0] == 0:
                salt = secrets.token_hex(16)
                pwd = "admin123"  # Cambiar en producción
                hash_pwd = hashlib.sha256(f"{pwd}{salt}".encode()).hexdigest()
                cursor.execute("""
                    INSERT INTO usuarios (username, password_hash, salt, nombre_completo, email, rol)
                    VALUES (?, ?, ?, 'Administrador', 'admin@satpro.com', 'admin')
                """, ('admin', f"{salt}${hash_pwd}", salt))
    
    def backup_database(self, automatic: bool = False):
        """Backup con compresión y rotación"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        prefix = "auto_" if automatic else "manual_"
        backup_name = f"{prefix}backup_{timestamp}.db"
        backup_path = os.path.join(config.BACKUP_DIR, backup_name)
        
        try:
            # Backup con SQLite
            with sqlite3.connect(config.DB_PATH) as src:
                with sqlite3.connect(backup_path) as dst:
                    src.backup(dst)
            
            # Comprimir
            compressed = f"{backup_path}.gz"
            import gzip
            with open(backup_path, 'rb') as f_in:
                with gzip.open(compressed, 'wb') as f_out:
                    shutil.copyfileobj(f_in, f_out)
            os.remove(backup_path)
            
            # Rotación: mantener solo últimos 30 backups automáticos
            if automatic:
                backups = sorted([f for f in os.listdir(config.BACKUP_DIR) if f.startswith('auto_')])
                for old in backups[:-30]:
                    os.remove(os.path.join(config.BACKUP_DIR, old))
            
            return compressed
        except Exception as e:
            st.error(f"Error en backup: {e}")
            return None
    
    def start_backup_scheduler(self):
        """Inicia scheduler de backups en thread separado"""
        def run_scheduler():
            schedule.every().day.at("02:00").do(self.backup_database, automatic=True)
            while True:
                schedule.run_pending()
                time.sleep(60)
        
        thread = threading.Thread(target=run_scheduler, daemon=True)
        thread.start()
    
    def close(self):
        """Cierra conexiones del thread actual"""
        if hasattr(self._local, 'conn') and self._local.conn:
            self._local.conn.close()
            self._local.conn = None

# Instancia global del gestor de base de datos
db = DatabaseManager()

# =============================================================================
# SISTEMA DE AUTENTICACIÓN ENTERPRISE
# =============================================================================

class AuthManager:
    """Sistema de autenticación enterprise con 2FA opcional"""
    
    def __init__(self):
        self.failed_attempts = {}
        self.lockout_duration = timedelta(minutes=30)
    
    def hash_password(self, password: str, salt: Optional[str] = None) -> Tuple[str, str]:
        """Hash seguro con salt"""
        if salt is None:
            salt = secrets.token_hex(16)
        salted = f"{password}{salt}{config.SECRET_KEY}"
        hash_obj = hashlib.sha256(salted.encode())
        return f"{salt}${hash_obj.hexdigest()}", salt
    
    def verify_password(self, password: str, stored_hash: str) -> bool:
        """Verificación de hash"""
        try:
            salt, _ = stored_hash.split('$')
            check_hash, _ = self.hash_password(password, salt)
            return check_hash == stored_hash
        except ValueError:
            return False
    
    def is_locked(self, username: str) -> bool:
        """Verifica si usuario está bloqueado"""
        with db.get_cursor() as cursor:
            cursor.execute("""
                SELECT bloqueado_hasta FROM usuarios WHERE username = ?
            """, (username,))
            result = cursor.fetchone()
            if result and result[0]:
                bloqueado = datetime.fromisoformat(result[0])
                return bloqueado > datetime.now()
        return False
    
    def record_attempt(self, username: str, success: bool):
        """Registra intento de login"""
        with db.get_cursor() as cursor:
            if success:
                cursor.execute("""
                    UPDATE usuarios 
                    SET intentos_fallidos = 0, 
                        bloqueado_hasta = NULL,
                        ultimo_acceso = CURRENT_TIMESTAMP
                    WHERE username = ?
                """, (username,))
            else:
                cursor.execute("""
                    UPDATE usuarios 
                    SET intentos_fallidos = intentos_fallidos + 1,
                        bloqueado_hasta = CASE 
                            WHEN intentos_fallidos >= 4 THEN datetime('now', '+30 minutes')
                            ELSE bloqueado_hasta 
                        END
                    WHERE username = ?
                """, (username,))
    
    def authenticate(self, username: str, password: str, ip: str = None) -> Optional[Dict]:
        """Autenticación completa"""
        if self.is_locked(username):
            return {"error": "Cuenta bloqueada temporalmente por intentos fallidos"}
        
        with db.get_cursor() as cursor:
            cursor.execute("""
                SELECT id, username, password_hash, nombre_completo, email, rol, activo 
                FROM usuarios WHERE username = ? AND activo = 1
            """, (username,))
            
            user = cursor.fetchone()
            if not user:
                self.record_attempt(username, False)
                return None
            
            if not self.verify_password(password, user['password_hash']):
                self.record_attempt(username, False)
                return None
            
            self.record_attempt(username, True)
            
            return {
                "id": user['id'],
                "username": user['username'],
                "nombre": user['nombre_completo'],
                "email": user['email'],
                "role": user['rol'],
                "ip": ip
            }
    
    def create_user(self, username: str, password: str, nombre: str, email: str, rol: str = 'tecnico'):
        """Crea nuevo usuario"""
        hash_pwd, salt = self.hash_password(password)
        
        with db.get_cursor() as cursor:
            try:
                cursor.execute("""
                    INSERT INTO usuarios (username, password_hash, salt, nombre_completo, email, rol)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (username, hash_pwd, salt, nombre, email, rol))
                return True
            except sqlite3.IntegrityError:
                return False

auth_manager = AuthManager()

# =============================================================================
# SISTEMA DE NOTIFICACIONES MULTICANAL
# =============================================================================

class NotificationManager:
    """Gestor de notificaciones Email, WhatsApp y Push"""
    
    def __init__(self):
        self.mp = mercadopago.SDK(config.MP_ACCESS_TOKEN) if config.MP_ACCESS_TOKEN else None
    
    def send_email(self, to: str, subject: str, body: str, html: bool = True) -> bool:
        """Envío de email vía SMTP"""
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = config.EMAIL_FROM
            msg['To'] = to
            
            if html:
                msg.attach(MIMEText(body, 'html'))
            else:
                msg.attach(MIMEText(body, 'plain'))
            
            with smtplib.SMTP(config.SMTP_SERVER, config.SMTP_PORT) as server:
                server.starttls()
                server.login(config.SMTP_USER, config.SMTP_PASS)
                server.send_message(msg)
            
            # Registrar en DB
            with db.get_cursor() as cursor:
                cursor.execute("""
                    INSERT INTO notificaciones (tipo, destinatario, asunto, contenido, estado, enviado_at)
                    VALUES (?, ?, ?, ?, 'enviado', CURRENT_TIMESTAMP)
                """, ('email', to, subject, body))
            
            return True
        except Exception as e:
            # Registrar error
            with db.get_cursor() as cursor:
                cursor.execute("""
                    INSERT INTO notificaciones (tipo, destinatario, asunto, contenido, estado, error_msg)
                    VALUES (?, ?, ?, ?, 'error', ?)
                """, ('email', to, subject, body, str(e)))
            return False
    
    def send_whatsapp(self, phone: str, message: str, orden_id: int = None) -> bool:
        """Envío de WhatsApp Business API"""
        if not config.WHATSAPP_API_KEY:
            return False
        
        try:
            # Normalizar número
            phone = re.sub(r'\D', '', phone)
            if not phone.startswith('54'):
                phone = '54' + phone
            
            headers = {
                'Authorization': f'Bearer {config.WHATSAPP_API_KEY}',
                'Content-Type': 'application/json'
            }
            
            payload = {
                'messaging_product': 'whatsapp',
                'recipient_type': 'individual',
                'to': phone,
                'type': 'text',
                'text': {'body': message}
            }
            
            response = requests.post(
                config.WHATSAPP_API_URL,
                headers=headers,
                json=payload,
                timeout=30
            )
            
            success = response.status_code == 200
            
            with db.get_cursor() as cursor:
                cursor.execute("""
                    INSERT INTO notificaciones (tipo, destinatario, contenido, estado, orden_id, enviado_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, ('whatsapp', phone, message, 'enviado' if success else 'error', 
                      orden_id, datetime.now() if success else None))
            
            return success
            
        except Exception as e:
            with db.get_cursor() as cursor:
                cursor.execute("""
                    INSERT INTO notificaciones (tipo, destinatario, contenido, estado, orden_id, error_msg)
                    VALUES (?, ?, ?, 'error', ?, ?)
                """, ('whatsapp', phone, message, orden_id, str(e)))
            return False
    
    def notify_status_change(self, orden_id: int, nuevo_estado: str, cliente_data: Dict):
        """Notifica cambio de estado al cliente"""
        mensajes = {
            'Diagnóstico': f"Su vehículo está siendo diagnosticado. Le informaremos el presupuesto en breve.",
            'Presupuestado': f"Presupuesto listo. Por favor apruebe para continuar con la reparación.",
            'En reparación': f"¡Buenas noticias! Su vehículo está en reparación.",
            'Listo para entrega': f"✅ Su vehículo está listo. Puede retirarlo en horario de atención.",
            'Entregado': f"Gracias por confiar en nosotros. ¿Todo funcionó correctamente con su vehículo?"
        }
        
        mensaje = mensajes.get(nuevo_estado, f"Estado actualizado a: {nuevo_estado}")
        
        # Email
        if cliente_data.get('email'):
            html_body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                        <h2 style="color: #1a73e8;">SAT Pro - Actualización de Servicio</h2>
                        <p>Estimado {cliente_data['nombre']},</p>
                        <p>{mensaje}</p>
                        <p><strong>Número de Orden:</strong> #{orden_id}</p>
                        <hr>
                        <p style="font-size: 12px; color: #666;">
                            Este es un mensaje automático. No responda a este email.<br>
                            Para consultas: {config.EMAIL_FROM}
                        </p>
                    </div>
                </body>
            </html>
            """
            self.send_email(cliente_data['email'], f"Orden #{orden_id} - {nuevo_estado}", html_body)
        
        # WhatsApp
        if cliente_data.get('telefono'):
            self.send_whatsapp(cliente_data['telefono'], 
                f"🔧 SAT Pro - Orden #{orden_id}\n{mensaje}\n\nNo responda este mensaje.",
                orden_id)

notification_manager = NotificationManager()

# =============================================================================
# INTEGRACIÓN MERCADOPAGO
# =============================================================================

class PaymentManager:
    """Gestor de pagos integrado"""
    
    def __init__(self):
        self.sdk = mercadopago.SDK(config.MP_ACCESS_TOKEN) if config.MP_ACCESS_TOKEN else None
    
    def create_preference(self, orden_id: int, items: List[Dict], cliente: Dict) -> Optional[str]:
        """Crea preferencia de pago en MercadoPago"""
        if not self.sdk:
            return None
        
        preference_data = {
            "items": [
                {
                    "title": item['descripcion'],
                    "quantity": item.get('cantidad', 1),
                    "unit_price": float(item['precio'])
                } for item in items
            ],
            "payer": {
                "name": cliente['nombre'],
                "email": cliente.get('email', ''),
                "phone": {"number": cliente.get('telefono', '')}
            },
            "external_reference": str(orden_id),
            "notification_url": "https://satpro.com/webhooks/mp",
            "back_urls": {
                "success": f"https://satpro.com/orden/{orden_id}/success",
                "failure": f"https://satpro.com/orden/{orden_id}/failure",
                "pending": f"https://satpro.com/orden/{orden_id}/pending"
            },
            "auto_return": "approved"
        }
        
        try:
            preference_response = self.sdk.preference().create(preference_data)
            return preference_response["response"]["init_point"]
        except Exception as e:
            st.error(f"Error creando preferencia MP: {e}")
            return None
    
    def process_webhook(self, data: Dict):
        """Procesa notificaciones de MP"""
        # Implementar lógica de webhook
        pass

payment_manager = PaymentManager()

# =============================================================================
# GENERADOR DE REPORTES EXCEL Y PDF
# =============================================================================

class ReportGenerator:
    """Generador de reportes empresariales"""
    
    @staticmethod
    def generate_excel_report(tipo: str, fecha_desde: datetime, fecha_hasta: datetime) -> BytesIO:
        """Genera reportes Excel formateados"""
        wb = Workbook()
        ws = wb.active
        
        # Estilos
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if tipo == "ordenes":
            ws.title = "Órdenes de Servicio"
            
            # Headers
            headers = ["N° Orden", "Fecha", "Cliente", "Vehículo", "Estado", "Total", "Método Pago"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Datos
            with db.get_cursor() as cursor:
                cursor.execute("""
                    SELECT r.numero_orden, r.fecha_ingreso, c.nombre, v.patente, 
                           r.estado, r.total, r.metodo_pago
                    FROM reparaciones r
                    JOIN clientes c ON r.cliente_id = c.id
                    LEFT JOIN vehiculos v ON r.vehiculo_id = v.id
                    WHERE DATE(r.fecha_ingreso) BETWEEN ? AND ?
                    ORDER BY r.fecha_ingreso DESC
                """, (fecha_desde.strftime('%Y-%m-%d'), fecha_hasta.strftime('%Y-%m-%d')))
                
                for row_idx, row in enumerate(cursor.fetchall(), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        if col_idx == 6:  # Total
                            cell.number_format = '$#,##0.00'
        
        elif tipo == "inventario":
            ws.title = "Inventario Valorizado"
            
            headers = ["Código", "Item", "Stock", "Mínimo", "Costo", "Venta", "Margen %", "Valor Stock"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            with db.get_cursor() as cursor:
                cursor.execute("""
                    SELECT codigo, item, cantidad, minimo, precio_costo, 
                           precio_venta, margen, (cantidad * precio_costo) as valor_stock
                    FROM inventario WHERE activo = 1 ORDER BY item
                """)
                
                for row_idx, row in enumerate(cursor.fetchall(), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        if col_idx in [5, 6, 8]:  # Monetarios
                            cell.number_format = '$#,##0.00'
                        elif col_idx == 7:  # Margen
                            cell.number_format = '0.00%'
        
        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_qr_for_orden(orden_id: int) -> str:
        """Genera QR para seguimiento de orden"""
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(f"https://satpro.com/seguimiento/{orden_id}")
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()
        return f"data:image/png;base64,{img_str}"

report_generator = ReportGenerator()

# =============================================================================
# DOCKER Y DEPLOYMENT
# =============================================================================

def generate_docker_files():
    """Genera archivos Docker para deployment"""
    
    dockerfile_content = """FROM python:3.11-slim

WORKDIR /app

# Instalar dependencias del sistema
RUN apt-get update && apt-get install -y \\
    gcc \\
    libsqlite3-dev \\
    && rm -rf /var/lib/apt/lists/*

# Copiar requirements
COPY requirements.txt .
RUN pip install --no-cache-dir -r requirements.txt

# Copiar aplicación
COPY . .

# Crear directorios necesarios
RUN mkdir -p /app/backups /app/temp /app/exports

# Puerto
EXPOSE 8501

# Healthcheck
HEALTHCHECK CMD curl --fail http://localhost:8501/_stcore/health || exit 1

# Comando
ENTRYPOINT ["streamlit", "run", "app.py", "--server.port=8501", "--server.address=0.0.0.0"]"""
    
    docker_compose_content = """version: '3.8'

services:
  satpro:
    build: .
    container_name: satpro_app
    ports:
      - "8501:8501"
    volumes:
      - ./data:/app/data
      - ./backups:/app/backups
      - ./exports:/app/exports
    environment:
      - DB_PATH=/app/data/gestion_taller.db
      - BACKUP_DIR=/app/backups
      - SECRET_KEY=${SECRET_KEY}
      - MP_ACCESS_TOKEN=${MP_ACCESS_TOKEN}
      - SMTP_USER=${SMTP_USER}
      - SMTP_PASS=${SMTP_PASS}
    restart: unless-stopped
    networks:
      - satpro_network

  # Opcional: Backup automático con cron
  backup:
    image: alpine:latest
    volumes:
      - ./data:/data
      - ./backups:/backups
    command: >
      sh -c "echo '0 2 * * * sqlite3 /data/gestion_taller.db \".backup /backups/auto_$(date +\\%Y\\%m\\%d_\\%H\\%M\\%S).db\"' | crontab - && crond -f"
    restart: unless-stopped

networks:
  satpro_network:
    driver: bridge"""
    
    requirements_content = """streamlit>=1.28.0
pandas>=2.0.0
fpdf>=1.7.2
mercadopago>=2.2.0
openpyxl>=3.1.0
qrcode>=7.4.2
Pillow>=10.0.0
schedule>=1.2.0
requests>=2.31.0"""
    
    with open('Dockerfile', 'w') as f:
        f.write(dockerfile_content)
    
    with open('docker-compose.yml', 'w') as f:
        f.write(docker_compose_content)
        
    with open('requirements.txt', 'w') as f:
        f.write(requirements_content)
    
    st.success("✅ Archivos Docker y requirements generados")

# =============================================================================
# UI COMPONENTS ENTERPRISE
# =============================================================================

def login_screen():
    """Pantalla de login enterprise"""
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("""
        <div style="text-align: center; padding: 40px 0;">
            <div style="font-size: 4rem; margin-bottom: 20px;">🔧</div>
            <h1 style="color: #1a73e8; font-weight: 700; margin-bottom: 10px;">SAT Pro Enterprise</h1>
            <p style="color: #666; font-size: 1.1rem;">Sistema Integral de Gestión de Taller</p>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("login_form"):
            username = st.text_input("👤 Usuario", placeholder="Ingrese su usuario")
            password = st.text_input("🔒 Contraseña", type="password")
            
            col_btn1, col_btn2 = st.columns([1, 1])
            with col_btn1:
                submitted = st.form_submit_button("Iniciar Sesión", use_container_width=True)
            with col_btn2:
                docker_btn = st.form_submit_button("🐳 Generar Docker", use_container_width=True)
            
            if docker_btn:
                generate_docker_files()
            
            if submitted:
                if not username or not password:
                    st.error("Complete todos los campos")
                    return
                
                # Obtener IP real (simulado para demo)
                client_ip = "127.0.0.1"
                
                result = auth_manager.authenticate(username, password, client_ip)
                
                if result and "error" not in result:
                    st.session_state["authenticated"] = True
                    st.session_state["user"] = result
                    st.session_state["login_time"] = datetime.now()
                    st.rerun()
                elif result and "error" in result:
                    st.error(result["error"])
                else:
                    st.error("Credenciales incorrectas")
                    st.caption("⚠️ Después de 5 intentos fallidos, la cuenta se bloqueará por 30 minutos")
        
        st.markdown("---")
        st.caption(f"v3.0 Enterprise | {datetime.now().year} SAT Pro")

def sidebar_menu():
    """Menú lateral enterprise"""
    with st.sidebar:
        user = st.session_state["user"]
        
        st.markdown(f"""
        <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    border-radius: 16px; color: white; margin-bottom: 20px;">
            <div style="font-size: 3rem; margin-bottom: 10px;">👤</div>
            <div style="font-weight: 600; font-size: 1.1rem;">{user['nombre']}</div>
            <div style="font-size: 0.8rem; opacity: 0.9; text-transform: uppercase; letter-spacing: 1px;">
                {user['role']}
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Verificar tiempo de sesión
        if "login_time" in st.session_state:
            elapsed = datetime.now() - st.session_state["login_time"]
            remaining = timedelta(minutes=config.SESSION_TIMEOUT) - elapsed
            
            if remaining.total_seconds() <= 0:
                st.error("⏰ Sesión expirada")
                st.session_state.clear()
                st.rerun()
            
            minutes_left = int(remaining.total_seconds() / 60)
            st.caption(f"⏱️ Sesión: {minutes_left} min restantes")
        
        st.divider()
        
        menu_items = {
            "📊 Dashboard": "dashboard",
            "📦 Inventario": "inventario",
            "👥 Clientes": "clientes",
            "📝 Nueva Orden": "nueva_orden",
            "🔧 Taller (Órdenes)": "taller",
            "💰 Facturación": "facturacion",
            "📈 Reportes": "reportes",
            "⚙️ Configuración": "config"
        }
        
        # Filtrar por permisos
        if user['role'] not in ['admin', 'supervisor']:
            menu_items.pop("⚙️ Configuración")
            menu_items.pop("📈 Reportes")
        
        selected = st.radio("Navegación", list(menu_items.keys()))
        
        st.divider()
        
        # Acciones rápidas
        st.caption("⚡ Acciones Rápidas")
        if st.button("🔍 Buscar Orden", use_container_width=True):
            st.session_state["quick_search"] = True
        
        if st.button("💾 Backup Ahora", use_container_width=True):
            backup_file = db.backup_database(automatic=False)
            if backup_file:
                st.success(f"✅ Backup guardado: {os.path.basename(backup_file)}")
        
        st.divider()
        
        if st.button("🚪 Cerrar Sesión", use_container_width=True, type="secondary"):
            st.session_state.clear()
            st.rerun()
        
        return menu_items[selected]

# =============================================================================
# MÓDULOS PRINCIPALES
# =============================================================================

def dashboard_module():
    """Dashboard ejecutivo con analytics"""
    st.markdown('<h1 class="main-header">📊 Dashboard Ejecutivo</h1>', unsafe_allow_html=True)
    
    # KPIs principales
    with db.connection:
        kpis = pd.read_sql_query("""
            SELECT 
                COUNT(CASE WHEN estado NOT IN ('Entregado', 'Cancelado') THEN 1 END) as activas,
                COUNT(CASE WHEN estado = 'Entregado' AND DATE(fecha_finalizacion) = DATE('now') THEN 1 END) as entregadas_hoy,
                COALESCE(SUM(CASE WHEN DATE(fecha_finalizacion) = DATE('now') AND estado = 'Entregado' THEN total END), 0) as ingresos_hoy,
                COUNT(CASE WHEN estado = 'Urgente' OR (julianday('now') - julianday(fecha_ingreso) > 5 AND estado != 'Entregado') THEN 1 END) as alertas,
                (SELECT COUNT(*) FROM inventario WHERE cantidad <= minimo) as stock_critico,
                (SELECT COUNT(*) FROM clientes WHERE DATE(ultima_visita) >= DATE('now', '-30 days')) as clientes_activos
            FROM reparaciones
            WHERE activo = 1
        """, db.connection)
    
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    
    metrics = [
        ("🔧 Órdenes Activas", int(kpis.iloc[0]['activas']), "#667eea"),
        ("✅ Entregas Hoy", int(kpis.iloc[0]['entregadas_hoy']), "#28a745"),
        ("💰 Ingresos Hoy", f"${kpis.iloc[0]['ingresos_hoy']:,.0f}", "#17a2b8"),
        ("🚨 Alertas", int(kpis.iloc[0]['alertas']), "#dc3545"),
        ("📦 Stock Crítico", int(kpis.iloc[0]['stock_critico']), "#ffc107"),
        ("👥 Clientes 30d", int(kpis.iloc[0]['clientes_activos']), "#6f42c1")
    ]
    
    for col, (label, value, color) in zip([col1, col2, col3, col4, col5, col6], metrics):
        with col:
            st.markdown(f"""
            <div style="background: {color}; padding: 20px; border-radius: 12px; color: white; text-align: center;">
                <div style="font-size: 1.8rem; font-weight: 700;">{value}</div>
                <div style="font-size: 0.85rem; opacity: 0.9;">{label}</div>
            </div>
            """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Gráficos y análisis
    col_left, col_right = st.columns([2, 1])
    
    with col_left:
        st.subheader("Tendencias de Ingresos")
        
        tab1, tab2 = st.tabs(["Últimos 7 días", "Últimos 30 días"])
        
        with tab1:
            with db.connection:
                trend = pd.read_sql_query("""
                    SELECT DATE(fecha_ingreso) as fecha,
                           COUNT(*) as ordenes,
                           COALESCE(SUM(CASE WHEN estado = 'Entregado' THEN total END), 0) as ingresos
                    FROM reparaciones
                    WHERE fecha_ingreso >= DATE('now', '-7 days')
                    GROUP BY DATE(fecha_ingreso)
                    ORDER BY fecha
                """, db.connection)
            
            if not trend.empty:
                st.line_chart(trend.set_index('fecha')[['ordenes', 'ingresos']])
            else:
                st.info("Sin datos suficientes")
        
        with tab2:
            with db.connection:
                monthly = pd.read_sql_query("""
                    SELECT strftime('%Y-%m', fecha_ingreso) as mes,
                           COUNT(*) as ordenes,
                           COALESCE(SUM(total), 0) as ingresos
                    FROM reparaciones
                    WHERE fecha_ingreso >= DATE('now', '-30 days')
                    GROUP BY mes
                    ORDER BY mes
                """, db.connection)
            
            if not monthly.empty:
                st.bar_chart(monthly.set_index('mes')[['ingresos']])
    
    with col_right:
        st.subheader("Órdenes por Estado")
        with db.connection:
            estados = pd.read_sql_query("""
                SELECT estado, COUNT(*) as cantidad,
                       ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 1) as porcentaje
                FROM reparaciones
                WHERE estado NOT IN ('Entregado', 'Cancelado') AND activo = 1
                GROUP BY estado
                ORDER BY cantidad DESC
            """, db.connection)
        
        for _, row in estados.iterrows():
            st.progress(row['porcentaje'] / 100, text=f"{row['estado']} ({row['cantidad']})")
        
        # Alertas operativas
        st.subheader("🚨 Alertas Operativas")
        
        with db.connection:
            alertas_df = pd.read_sql_query("""
                SELECT r.id, r.cliente_id, c.nombre as cliente, r.equipo, r.estado,
                       julianday('now') - julianday(r.fecha_ingreso) as dias
                FROM reparaciones r
                JOIN clientes c ON r.cliente_id = c.id
                WHERE (r.estado = 'Urgente' OR (julianday('now') - julianday(r.fecha_ingreso) > 5))
                AND r.estado != 'Entregado'
                ORDER BY dias DESC
                LIMIT 5
            """, db.connection)
        
        if not alertas_df.empty:
            for _, alerta in alertas_df.iterrows():
                st.error(f"⚠️ Orden #{alerta['id']} - {alerta['cliente']} ({int(alerta['dias'])} días en {alerta['estado']})")
        else:
            st.success("✅ Sin alertas operativas")

def inventario_module():
    """Gestión completa de inventario"""
    st.markdown('<h1 class="main-header">📦 Gestión de Inventario</h1>', unsafe_allow_html=True)
    
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Stock", "➕ Nuevo Item", "🛒 Órdenes de Compra", "📊 Valorización"])
    
    with tab1:
        with db.connection:
            inventario = pd.read_sql_query("""
                SELECT i.*, p.nombre as proveedor_nombre,
                       CASE 
                           WHEN i.cantidad = 0 THEN 'AGOTADO'
                           WHEN i.cantidad <= i.minimo THEN 'CRITICO'
                           WHEN i.cantidad <= i.minimo * 1.5 THEN 'BAJO'
                           ELSE 'OK'
                       END as estado_stock
                FROM inventario i
                LEFT JOIN proveedores p ON i.proveedor_id = p.id
                WHERE i.activo = 1
                ORDER BY 
                    CASE estado_stock
                        WHEN 'AGOTADO' THEN 1
                        WHEN 'CRITICO' THEN 2
                        WHEN 'BAJO' THEN 3
                        ELSE 4
                    END,
                    i.item
            """, db.connection)
        
        # Filtros avanzados
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            filtro_texto = st.text_input("🔍 Buscar")
        with col_f2:
            filtro_categoria = st.selectbox("Categoría", ["Todas", "Repuestos", "Insumos", "Accesorios"])
        with col_f3:
            filtro_estado = st.multiselect("Estado Stock", ["AGOTADO", "CRITICO", "BAJO", "OK"], default=[])
        
        df_filtered = inventario.copy()
        if filtro_texto:
            mask = df_filtered['item'].str.contains(filtro_texto, case=False) | \
                   df_filtered['codigo'].str.contains(filtro_texto, case=False)
            df_filtered = df_filtered[mask]
        if filtro_estado:
            df_filtered = df_filtered[df_filtered['estado_stock'].isin(filtro_estado)]
        
        # Mostrar con color
        st.dataframe(
            df_filtered,
            column_config={
                "precio_venta": st.column_config.NumberColumn("Precio Venta", format="$%.2f"),
                "precio_costo": st.column_config.NumberColumn("Precio Costo", format="$%.2f"),
                "margen": st.column_config.NumberColumn("Margen", format="%.1f%%"),
                "estado_stock": st.column_config.Column("Estado")
            },
            use_container_width=True,
            height=400
        )
        
        # Acciones masivas
        st.subheader("⚡ Acciones Rápidas")
        col_a1, col_a2, col_a3 = st.columns(3)
        
        with col_a1:
            if st.button("📋 Generar Lista de Compra", use_container_width=True):
                critico = df_filtered[df_filtered['estado_stock'].isin(['AGOTADO', 'CRITICO'])]
                if not critico.empty:
                    lista = "ORDEN DE COMPRA AUTOMÁTICA\n"
                    lista += f"Fecha: {datetime.now().strftime('%d/%m/%Y')}\n\n"
                    for _, item in critico.iterrows():
                        faltante = max(0, item['minimo'] * 2 - item['cantidad'])
                        lista += f"□ {item['codigo']} - {item['item']}\n"
                        lista += f"  Cantidad sugerida: {faltante} | Proveedor: {item['proveedor_nombre'] or 'N/A'}\n"
                        lista += f"  Precio referencia: ${item['precio_costo'] or 0:.2f}\n\n"
                    
                    st.download_button("⬇️ Descargar OC", lista, 
                                     file_name=f"OC_Generada_{datetime.now().strftime('%Y%m%d')}.txt")
        
        with col_a2:
            if st.button("📊 Exportar a Excel", use_container_width=True):
                excel_file = report_generator.generate_excel_report("inventario", 
                                                                datetime.now() - timedelta(days=365), 
                                                                datetime.now())
                st.download_button("⬇️ Descargar Excel", excel_file, 
                                 file_name=f"Inventario_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        with col_a3:
            if st.button("🔄 Actualizar Precios Masivo", use_container_width=True):
                st.info("Función: Subir CSV con códigos y nuevos precios")
    
    with tab2:
        with st.form("nuevo_item"):
            col1, col2 = st.columns(2)
            with col1:
                codigo = st.text_input("Código SKU *", placeholder="REP-001")
                nombre = st.text_input("Nombre *", placeholder="Filtro de Aceite")
                categoria = st.selectbox("Categoría", ["Repuestos", "Insumos", "Accesorios", "Herramientas"])
                cantidad = st.number_input("Stock Inicial", min_value=0, value=0)
            with col2:
                minimo = st.number_input("Stock Mínimo", min_value=1, value=5)
                costo = st.number_input("Precio Costo", min_value=0.0, value=0.0)
                venta = st.number_input("Precio Venta *", min_value=0.0, value=0.0)
                proveedor = st.selectbox("Proveedor", ["Sin proveedor"])  # Cargar dinámicamente
            
            ubicacion = st.text_input("Ubicación en Depósito", placeholder="Estante A3 - Fila 2")
            descripcion = st.text_area("Descripción Técnica")
            
            submitted = st.form_submit_button("💾 Guardar Item")
            if submitted:
                if not all([codigo, nombre, venta > 0]):
                    st.error("Complete campos obligatorios")
                else:
                    try:
                        with db.get_cursor() as cursor:
                            cursor.execute("""
                                INSERT INTO inventario 
                                (codigo, item, categoria, descripcion, cantidad, minimo, 
                                 precio_costo, precio_venta, ubicacion)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (codigo.upper(), nombre, categoria, descripcion, cantidad, 
                                  minimo, costo, venta, ubicacion))
                        st.success(f"✅ {nombre} guardado correctamente")
                    except sqlite3.IntegrityError:
                        st.error(f"❌ Código {codigo} ya existe")
    
    with tab3:
        st.subheader("🛒 Gestión de Compras")
        
        with db.connection:
            proveedores = pd.read_sql_query("SELECT id, nombre FROM proveedores WHERE activo = 1", db.connection)
        
        if not proveedores.empty:
            prov_sel = st.selectbox("Proveedor", proveedores['nombre'].tolist())
            
            # Simulador de orden de compra
            st.write("**Items a solicitar:**")
            # Aquí iría un formulario dinámico para agregar items
            
            if st.button("📤 Enviar OC por Email", use_container_width=True):
                st.success("Orden de compra enviada al proveedor")
    
    with tab4:
        st.subheader("📊 Valorización de Inventario")
        
        with db.connection:
            valorizacion = pd.read_sql_query("""
                SELECT 
                    categoria,
                    COUNT(*) as items,
                    SUM(cantidad * precio_costo) as valor_costo,
                    SUM(cantidad * precio_venta) as valor_venta,
                    SUM(cantidad * (precio_venta - precio_costo)) as margen_potencial
                FROM inventario
                WHERE activo = 1
                GROUP BY categoria
            """, db.connection)
        
        if not valorizacion.empty:
            col_v1, col_v2, col_v3 = st.columns(3)
            with col_v1:
                st.metric("Valor al Costo", f"${valorizacion['valor_costo'].sum():,.2f}")
            with col_v2:
                st.metric("Valor de Venta", f"${valorizacion['valor_venta'].sum():,.2f}")
            with col_v3:
                st.metric("Margen Potencial", f"${valorizacion['margen_potencial'].sum():,.2f}")
            
            st.bar_chart(valorizacion.set_index('categoria')[['valor_costo', 'valor_venta']])

def nueva_orden_module():
    """Creación de orden con workflow completo"""
    st.markdown('<h1 class="main-header">📝 Nueva Orden de Servicio</h1>', unsafe_allow_html=True)
    
    # Wizard de pasos
    step = st.session_state.get("orden_step", 1)
    
    col_steps = st.columns(4)
    steps = ["1. Cliente", "2. Vehículo", "3. Diagnóstico", "4. Confirmación"]
    for i, (col, label) in enumerate(zip(col_steps, steps), 1):
        with col:
            if i == step:
                st.markdown(f"<div style='background: #1a73e8; color: white; padding: 10px; border-radius: 8px; text-align: center; font-weight: 600;'>{label}</div>", unsafe_allow_html=True)
            elif i < step:
                st.markdown(f"<div style='background: #28a745; color: white; padding: 10px; border-radius: 8px; text-align: center;'>✅ {label}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #e9ecef; color: #6c757d; padding: 10px; border-radius: 8px; text-align: center;'>{label}</div>", unsafe_allow_html=True)
    
    # Paso 1: Selección/Creación de Cliente
    if step == 1:
        st.subheader("👥 Seleccionar o Crear Cliente")
        
        with db.connection:
            clientes = pd.read_sql_query("""
                SELECT id, nombre, apellido, telefono, email, dni_cuit 
                FROM clientes WHERE activo = 1 ORDER BY nombre
            """, db.connection)
        
        tab_buscar, tab_nuevo = st.tabs(["🔍 Buscar Existente", "➕ Nuevo Cliente"])
        
        with tab_buscar:
            if not clientes.empty:
                busqueda = st.text_input("Buscar por nombre, DNI o teléfono")
                if busqueda:
                    filtrados = clientes[
                        clientes['nombre'].str.contains(busqueda, case=False) |
                        clientes['apellido'].str.contains(busqueda, case=False) |
                        clientes['telefono'].str.contains(busqueda) |
                        clientes['dni_cuit'].str.contains(busqueda)
                    ]
                else:
                    filtrados = clientes
                
                for _, cli in filtrados.head(5).iterrows():
                    with st.container():
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            st.write(f"**{cli['nombre']} {cli['apellido']}**")
                            st.caption(f"📱 {cli['telefono']} | 🆔 {cli['dni_cuit']}")
                        with col2:
                            if st.button("Seleccionar", key=f"sel_cli_{cli['id']}"):
                                st.session_state["cliente_id"] = cli['id']
                                st.session_state["cliente_data"] = cli.to_dict()
                                st.session_state["orden_step"] = 2
                                st.rerun()
                        st.divider()
            else:
                st.info("No hay clientes registrados. Cree uno nuevo.")
        
        with tab_nuevo:
            with st.form("nuevo_cliente"):
                col1, col2 = st.columns(2)
                with col1:
                    nombre = st.text_input("Nombre *")
                    apellido = st.text_input("Apellido")
                    telefono = st.text_input("Teléfono *")
                with col2:
                    email = st.text_input("Email")
                    dni = st.text_input("DNI/CUIT")
                    direccion = st.text_input("Dirección")
                
                if st.form_submit_button("💾 Crear y Continuar"):
                    if nombre and telefono:
                        with db.get_cursor() as cursor:
                            cursor.execute("""
                                INSERT INTO clientes (nombre, apellido, telefono, email, dni_cuit, direccion)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (nombre, apellido, telefono, email, dni, direccion))
                            new_id = cursor.lastrowid
                        
                        st.session_state["cliente_id"] = new_id
                        st.session_state["cliente_data"] = {
                            "id": new_id, "nombre": nombre, "apellido": apellido,
                            "telefono": telefono, "email": email
                        }
                        st.session_state["orden_step"] = 2
                        st.rerun()
    
    # Paso 2: Vehículo
    elif step == 2:
        st.subheader("🚗 Seleccionar o Agregar Vehículo")
        cliente_id = st.session_state.get("cliente_id")
        
        with db.connection:
            vehiculos = pd.read_sql_query("""
                SELECT * FROM vehiculos WHERE cliente_id = ? AND activo = 1
            """, db.connection, params=(cliente_id,))
        
        if not vehiculos.empty:
            st.write("**Vehículos del cliente:**")
            for _, veh in vehiculos.iterrows():
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"**{veh['marca']} {veh['modelo']} ({veh['anio']})**")
                    st.caption(f"🚗 Patente: {veh['patente']} | 🛣️ {veh['kilometraje_actual']} km")
                with col2:
                    if st.button("Seleccionar", key=f"sel_veh_{veh['id']}"):
                        st.session_state["vehiculo_id"] = veh['id']
                        st.session_state["vehiculo_data"] = veh.to_dict()
                        st.session_state["orden_step"] = 3
                        st.rerun()
        
        with st.expander("➕ Agregar Nuevo Vehículo"):
            with st.form("nuevo_vehiculo"):
                col1, col2 = st.columns(2)
                with col1:
                    patente = st.text_input("Patente *").upper()
                    marca = st.text_input("Marca *")
                    modelo = st.text_input("Modelo *")
                with col2:
                    anio = st.number_input("Año", min_value=1900, max_value=datetime.now().year, value=2020)
                    color = st.text_input("Color")
                    km = st.number_input("Kilometraje", min_value=0, value=0)
                
                if st.form_submit_button("Guardar Vehículo"):
                    if all([patente, marca, modelo]):
                        with db.get_cursor() as cursor:
                            cursor.execute("""
                                INSERT INTO vehiculos (cliente_id, patente, marca, modelo, anio, color, kilometraje_actual)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (cliente_id, patente, marca, modelo, anio, color, km))
                            new_veh_id = cursor.lastrowid
                        
                        st.session_state["vehiculo_id"] = new_veh_id
                        st.session_state["vehiculo_data"] = {
                            "id": new_veh_id, "patente": patente, "marca": marca,
                            "modelo": modelo, "anio": anio
                        }
                        st.session_state["orden_step"] = 3
                        st.rerun()
    
    # Paso 3: Diagnóstico y Presupuesto
    elif step == 3:
        st.subheader("🔧 Diagnóstico y Presupuesto")
        
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**Cliente:** {st.session_state['cliente_data']['nombre']}")
            st.write(f"**Vehículo:** {st.session_state['vehiculo_data']['marca']} {st.session_state['vehiculo_data']['modelo']}")
        
        with col2:
            prioridad = st.selectbox("Prioridad", ["Baja", "Normal", "Alta", "Urgente"])
            fecha_estimada = st.date_input("Fecha Estimada Entrega", 
                                          value=datetime.now() + timedelta(days=3))
        
        falla = st.text_area("Descripción de la Falla *", height=100,
                           placeholder="Describa los síntomas reportados por el cliente...")
        
        diagnostico = st.text_area("Diagnóstico Técnico", height=100,
                                  placeholder="Evaluación preliminar del técnico...")
        
        st.subheader("💰 Presupuesto")
        
        # Selector de repuestos
        with db.connection:
            stock = pd.read_sql_query("""
                SELECT id, codigo, item, precio_venta, cantidad 
                FROM inventario WHERE cantidad > 0 AND activo = 1
            """, db.connection)
        
        items_presupuesto = []
        if not stock.empty:
            seleccionados = st.multiselect("Repuestos sugeridos", 
                                          stock['item'].tolist(),
                                          help="Seleccione los items para el presupuesto")
            
            if seleccionados:
                cols = st.columns(min(len(seleccionados), 3))
                for idx, item in enumerate(seleccionados):
                    with cols[idx % 3]:
                        item_data = stock[stock['item'] == item].iloc[0]
                        cant = st.number_input(f"Cant. {item[:20]}", 
                                             min_value=1, 
                                             max_value=int(item_data['cantidad']),
                                             value=1,
                                             key=f"pres_cant_{item_data['id']}")
                        items_presupuesto.append({
                            "id": int(item_data['id']),
                            "item": item,
                            "cantidad": cant,
                            "precio_unitario": float(item_data['precio_venta']),
                            "total": float(item_data['precio_venta']) * cant
                        })
        
        mano_obra = st.number_input("Mano de Obra ($)", min_value=0.0, value=0.0, step=500.0)
        
        # Cálculo de totales
        total_repuestos = sum(i['total'] for i in items_presupuesto)
        subtotal = total_repuestos + mano_obra
        
        col_t1, col_t2, col_t3 = st.columns(3)
        with col_t1:
            st.metric("Repuestos", f"${total_repuestos:,.2f}")
        with col_t2:
            st.metric("Mano de Obra", f"${mano_obra:,.2f}")
        with col_t3:
            st.metric("TOTAL", f"${subtotal:,.2f}", delta_color="inverse")
        
        if st.button("✅ Guardar Orden", type="primary", use_container_width=True):
            # Generar número de orden único
            numero_orden = f"SAT-{datetime.now().strftime('%Y%m%d')}-{secrets.token_hex(3).upper()}"
            
            with db.get_cursor() as cursor:
                cursor.execute("""
                    INSERT INTO reparaciones 
                    (numero_orden, cliente_id, vehiculo_id, falla, diagnostico, 
                     prioridad, fecha_estimada_entrega, mano_obra, subtotal, total, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Recibido')
                """, (numero_orden, st.session_state['cliente_id'], 
                      st.session_state['vehiculo_id'], falla, diagnostico,
                      prioridad, fecha_estimada.strftime('%Y-%m-%d'),
                      mano_obra, subtotal, subtotal))
                
                orden_id = cursor.lastrowid
                
                # Guardar items presupuestados (no consumidos aún)
                for item in items_presupuesto:
                    cursor.execute("""
                        INSERT INTO historial_consumo 
                        (id_reparacion, id_repuesto, cantidad, precio_venta, usuario_id, fecha)
                        VALUES (?, ?, ?, ?, ?, datetime('now', '+1 day'))
                    """, (orden_id, item['id'], item['cantidad'], item['precio_unitario'],
                          st.session_state['user']['id']))
                
                # Registrar estado inicial
                cursor.execute("""
                    INSERT INTO historial_estados 
                    (id_reparacion, estado_nuevo, usuario_id, nota_cambio)
                    VALUES (?, 'Recibido', ?, 'Orden creada desde wizard')
                """, (orden_id, st.session_state['user']['id']))
            
            # Notificar al cliente
            cliente = st.session_state['cliente_data']
            if cliente.get('email') or cliente.get('telefono'):
                notification_manager.notify_status_change(
                    orden_id, 'Recibido', cliente
                )
            
            st.success(f"✅ Orden #{numero_orden} creada exitosamente")
            
            # QR para seguimiento
            qr_code = report_generator.generate_qr_for_orden(orden_id)
            st.image(qr_code, caption="Escaneé para seguimiento online", width=150)
            
            if st.button("🖨️ Imprimir Comprobante"):
                st.info("Generando PDF...")
            
            # Reset wizard
            for key in ['orden_step', 'cliente_id', 'vehiculo_id', 'cliente_data', 'vehiculo_data']:
                if key in st.session_state:
                    del st.session_state[key]
            
            st.balloons()
def taller_module():
    """Gestión de órdenes en taller"""
    st.markdown('<h1 class="main-header">🔧 Gestión de Taller</h1>', unsafe_allow_html=True)
    
    # Filtros
    col1, col2, col3 = st.columns(3)
    with col1:
        filtro_estado = st.multiselect("Estado", 
            ['Recibido', 'Diagnóstico', 'Presupuestado', 'Aprobado', 
             'Esperando repuestos', 'En reparación', 'Pruebas QC', 
             'Listo para entrega', 'Entregado', 'Garantía', 'Cancelado'],
            default=['Recibido', 'Diagnóstico', 'En reparación', 'Pruebas QC'])
    with col2:
        filtro_prioridad = st.multiselect("Prioridad", ['Baja', 'Normal', 'Alta', 'Urgente'], default=['Normal', 'Alta', 'Urgente'])
    with col3:
        busqueda = st.text_input("🔍 Buscar orden, cliente o patente")
    
    # Construir query dinámica
    query = """
        SELECT r.*, c.nombre as cliente_nombre, c.telefono as cliente_telefono,
               v.patente, v.marca, v.modelo, v.anio
        FROM reparaciones r
        JOIN clientes c ON r.cliente_id = c.id
        LEFT JOIN vehiculos v ON r.vehiculo_id = v.id
        WHERE r.activo = 1
    """
    params = []
    
    if filtro_estado:
        placeholders = ','.join(['?' for _ in filtro_estado])
        query += f" AND r.estado IN ({placeholders})"
        params.extend(filtro_estado)
    
    if filtro_prioridad:
        placeholders = ','.join(['?' for _ in filtro_prioridad])
        query += f" AND r.prioridad IN ({placeholders})"
        params.extend(filtro_prioridad)
    
    if busqueda:
        query += " AND (r.numero_orden LIKE ? OR c.nombre LIKE ? OR v.patente LIKE ?)"
        search_term = f"%{busqueda}%"
        params.extend([search_term, search_term, search_term])
    
    query += " ORDER BY CASE r.prioridad WHEN 'Urgente' THEN 1 WHEN 'Alta' THEN 2 WHEN 'Normal' THEN 3 ELSE 4 END, r.fecha_ingreso DESC"
    
    with db.connection:
        ordenes = pd.read_sql_query(query, db.connection, params=params)
    
    if ordenes.empty:
        st.info("No hay órdenes que coincidan con los filtros")
        return
    
    # Mostrar órdenes en cards
    for _, orden in ordenes.iterrows():
        # Color según estado
        color_estado = {
            'Recibido': '#6c757d',
            'Diagnóstico': '#17a2b8',
            'Presupuestado': '#ffc107',
            'Aprobado': '#28a745',
            'Esperando repuestos': '#fd7e14',
            'En reparación': '#007bff',
            'Pruebas QC': '#6f42c1',
            'Listo para entrega': '#20c997',
            'Entregado': '#28a745',
            'Garantía': '#e83e8c',
            'Cancelado': '#dc3545'
        }.get(orden['estado'], '#6c757d')
        
        with st.expander(f"#{orden['numero_orden']} | {orden['patente']} | {orden['cliente_nombre']} | {orden['estado']}", expanded=False):
            col_info, col_acciones = st.columns([3, 1])
            
            with col_info:
                st.markdown(f"""
                <div style="border-left: 4px solid {color_estado}; padding-left: 15px;">
                    <h4>{orden['marca']} {orden['modelo']} ({orden['anio']})</h4>
                    <p><strong>Cliente:</strong> {orden['cliente_nombre']} | 📱 {orden['cliente_telefono']}</p>
                    <p><strong>Falla:</strong> {orden['falla']}</p>
                    <p><strong>Diagnóstico:</strong> {orden['diagnostico'] or 'Pendiente'}</p>
                    <p><strong>Presupuesto:</strong> ${orden['total']:,.2f}</p>
                    <p><strong>Ingreso:</strong> {orden['fecha_ingreso']} | <strong>Estimada:</strong> {orden['fecha_estimada']}</p>
                </div>
                """, unsafe_allow_html=True)
                
                # Historial de estados
                with db.connection:
                    historial = pd.read_sql_query("""
                        SELECT he.*, u.nombre_completo as usuario_nombre
                        FROM historial_estados he
                        LEFT JOIN usuarios u ON he.usuario_id = u.id
                        WHERE he.id_reparacion = ?
                        ORDER BY he.fecha DESC
                    """, db.connection, params=(orden['id'],))
                
                if not historial.empty:
                    with st.expander("📋 Historial de cambios"):
                        for _, hist in historial.iterrows():
                            st.caption(f"{hist['fecha']}: {hist['estado_anterior'] or 'Inicio'} → {hist['estado_nuevo']} por {hist['usuario_nombre']}")
            
            with col_acciones:
                st.subheader("Acciones")
                
                # Cambiar estado
                nuevo_estado = st.selectbox(
                    "Nuevo estado",
                    ['Recibido', 'Diagnóstico', 'Presupuestado', 'Aprobado', 
                     'Esperando repuestos', 'En reparación', 'Pruebas QC', 
                     'Listo para entrega', 'Entregado', 'Garantía', 'Cancelado'],
                    index=['Recibido', 'Diagnóstico', 'Presupuestado', 'Aprobado', 
                           'Esperando repuestos', 'En reparación', 'Pruebas QC', 
                           'Listo para entrega', 'Entregado', 'Garantía', 'Cancelado'].index(orden['estado']) if orden['estado'] in ['Recibido', 'Diagnóstico', 'Presupuestado', 'Aprobado', 'Esperando repuestos', 'En reparación', 'Pruebas QC', 'Listo para entrega', 'Entregado', 'Garantía', 'Cancelado'] else 0,
                    key=f"estado_{orden['id']}"
                )
                
                nota_cambio = st.text_area("Nota del cambio", key=f"nota_{orden['id']}", height=80)
                
                if st.button("🔄 Actualizar Estado", key=f"btn_update_{orden['id']}", use_container_width=True):
                    with db.get_cursor() as cursor:
                        # Actualizar estado
                        cursor.execute("""
                            UPDATE reparaciones 
                            SET estado = ?, fecha_finalizacion = CASE WHEN ? = 'Entregado' THEN CURRENT_TIMESTAMP ELSE fecha_finalizacion END
                            WHERE id = ?
                        """, (nuevo_estado, nuevo_estado, orden['id']))
                        
                        # Registrar en historial
                        cursor.execute("""
                            INSERT INTO historial_estados (id_reparacion, estado_anterior, estado_nuevo, usuario_id, nota_cambio, ip_address)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (orden['id'], orden['estado'], nuevo_estado, st.session_state['user']['id'], nota_cambio, '127.0.0.1'))
                    
                    # Notificar al cliente
                    cliente_data = {
                        'nombre': orden['cliente_nombre'],
                        'email': None,  # Podrías cargar el email real
                        'telefono': orden['cliente_telefono']
                    }
                    notification_manager.notify_status_change(orden['id'], nuevo_estado, cliente_data)
                    
                    st.success(f"✅ Estado actualizado a {nuevo_estado}")
                    st.rerun()
                
                # Acciones adicionales según estado
                if orden['estado'] == 'Listo para entrega':
                    if st.button("💰 Procesar Pago", key=f"pago_{orden['id']}", use_container_width=True, type="primary"):
                        st.session_state['orden_pago_id'] = orden['id']
                        st.rerun()
                
                if orden['estado'] == 'Entregado' and not orden['pagado']:
                    st.warning("⚠️ Pendiente de pago")
                
                # Ver detalles completos
                with st.expander("🔍 Ver detalles completos"):
                    st.json(orden.to_dict())

def facturacion_module():
    """Módulo de facturación y pagos"""
    st.markdown('<h1 class="main-header">💰 Facturación y Pagos</h1>', unsafe_allow_html=True)
    
    # Órdenes listas para facturar
    with db.connection:
        ordenes_pendientes = pd.read_sql_query("""
            SELECT r.*, c.nombre as cliente_nombre, c.email as cliente_email, c.telefono as cliente_telefono,
                   v.patente, v.marca, v.modelo
            FROM reparaciones r
            JOIN clientes c ON r.cliente_id = c.id
            LEFT JOIN vehiculos v ON r.vehiculo_id = v.id
            WHERE r.estado = 'Listo para entrega' AND r.pagado = 0 AND r.activo = 1
            ORDER BY r.fecha_ingreso
        """, db.connection)
    
    if ordenes_pendientes.empty:
        st.success("✅ No hay órdenes pendientes de facturación")
    else:
        st.subheader(f"📋 Órdenes Listas para Entrega ({len(ordenes_pendientes)})")
        
        for _, orden in ordenes_pendientes.iterrows():
            with st.container():
                col1, col2, col3 = st.columns([2, 2, 1])
                
                with col1:
                    st.write(f"**#{orden['numero_orden']}**")
                    st.write(f"{orden['cliente_nombre']}")
                    st.caption(f"{orden['marca']} {orden['modelo']} - {orden['patente']}")
                
                with col2:
                    st.write(f"**Total: ${orden['total']:,.2f}**")
                    st.caption(f"Mano de obra: ${orden['mano_obra']:,.2f}")
                
                with col3:
                    if st.button("💳 Cobrar", key=f"cobrar_{orden['id']}", use_container_width=True):
                        st.session_state['orden_cobro_id'] = orden['id']
                        st.rerun()
                
                st.divider()
    
    # Proceso de cobro
    if 'orden_cobro_id' in st.session_state:
        orden_id = st.session_state['orden_cobro_id']
        
        with db.connection:
            orden = pd.read_sql_query("""
                SELECT r.*, c.nombre as cliente_nombre, c.email as cliente_email
                FROM reparaciones r
                JOIN clientes c ON r.cliente_id = c.id
                WHERE r.id = ?
            """, db.connection, params=(orden_id,)).iloc[0]
        
        st.subheader("💳 Procesar Cobro")
        
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**Orden:** #{orden['numero_orden']}")
            st.write(f"**Cliente:** {orden['cliente_nombre']}")
            st.write(f"**Total:** ${orden['total']:,.2f}")
        
        with col2:
            metodo_pago = st.selectbox("Método de pago", 
                ['efectivo', 'tarjeta', 'transferencia', 'mercadopago', 'cheque', 'cuenta_corriente'])
            
            if metodo_pago == 'mercadopago':
                if st.button("🔗 Generar Link de Pago MercadoPago", use_container_width=True):
                    items = [{
                        'descripcion': f'Servicio Taller - Orden {orden["numero_orden"]}',
                        'cantidad': 1,
                        'precio': float(orden['total'])
                    }]
                    cliente = {
                        'nombre': orden['cliente_nombre'],
                        'email': orden['cliente_email'],
                        'telefono': ''
                    }
                    link = payment_manager.create_preference(orden_id, items, cliente)
                    if link:
                        st.success("✅ Link generado")
                        st.markdown(f"[**Click aquí para pagar**]({link})")
                        st.code(link)
        
        descuento = st.number_input("Descuento ($)", min_value=0.0, max_value=float(orden['total']), value=0.0, step=50.0)
        total_final = orden['total'] - descuento
        
        st.write(f"**Total con descuento:** ${total_final:,.2f}")
        
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("✅ Confirmar Cobro", use_container_width=True, type="primary"):
                with db.get_cursor() as cursor:
                    # Registrar pago
                    cursor.execute("""
                        INSERT INTO pagos (orden_id, monto, metodo, referencia, estado)
                        VALUES (?, ?, ?, ?, 'completado')
                    """, (orden_id, total_final, metodo_pago, f'PAGO-{datetime.now().strftime("%Y%m%d%H%M%S")}'))
                    
                    # Marcar orden como pagada
                    cursor.execute("""
                        UPDATE reparaciones 
                        SET pagado = 1, metodo_pago = ?, descuento_monto = ?, total = ?
                        WHERE id = ?
                    """, (metodo_pago, descuento, total_final, orden_id))
                
                st.success("✅ Cobro registrado correctamente")
                del st.session_state['orden_cobro_id']
                st.rerun()
        
        with col_btn2:
            if st.button("❌ Cancelar", use_container_width=True):
                del st.session_state['orden_cobro_id']
                st.rerun()

def reportes_module():
    """Módulo de reportes y estadísticas"""
    st.markdown('<h1 class="main-header">📈 Reportes y Estadísticas</h1>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["📊 Reportes de Ventas", "📦 Reportes de Inventario", "👥 Reportes de Clientes"])
    
    with tab1:
        st.subheader("Reportes de Ventas")
        
        col1, col2 = st.columns(2)
        with col1:
            fecha_desde = st.date_input("Desde", value=datetime.now() - timedelta(days=30))
        with col2:
            fecha_hasta = st.date_input("Hasta", value=datetime.now())
        
        if st.button("📥 Generar Reporte de Órdenes", use_container_width=True):
            excel_file = report_generator.generate_excel_report("ordenes", fecha_desde, fecha_hasta)
            st.download_button(
                "⬇️ Descargar Excel",
                excel_file,
                file_name=f"Reporte_Ordenes_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Estadísticas rápidas
        with db.connection:
            stats = pd.read_sql_query("""
                SELECT 
                    COUNT(*) as total_ordenes,
                    SUM(CASE WHEN estado = 'Entregado' THEN 1 ELSE 0 END) as completadas,
                    SUM(CASE WHEN estado = 'Entregado' THEN total ELSE 0 END) as ingresos,
                    AVG(CASE WHEN estado = 'Entregado' THEN total ELSE NULL END) as ticket_promedio
                FROM reparaciones
                WHERE DATE(fecha_ingreso) BETWEEN ? AND ?
            """, db.connection, params=(fecha_desde.strftime('%Y-%m-%d'), fecha_hasta.strftime('%Y-%m-%d')))
        
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        with col_stat1:
            st.metric("Total Órdenes", int(stats.iloc[0]['total_ordenes']))
        with col_stat2:
            st.metric("Completadas", int(stats.iloc[0]['completadas']))
        with col_stat3:
            st.metric("Ingresos", f"${stats.iloc[0]['ingresos'] or 0:,.2f}")
        with col_stat4:
            st.metric("Ticket Promedio", f"${stats.iloc[0]['ticket_promedio'] or 0:,.2f}")
    
    with tab2:
        st.subheader("Reportes de Inventario")
        
        if st.button("📥 Generar Reporte de Inventario Actual", use_container_width=True):
            excel_file = report_generator.generate_excel_report("inventario", datetime.now() - timedelta(days=365), datetime.now())
            st.download_button(
                "⬇️ Descargar Excel",
                excel_file,
                file_name=f"Inventario_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Movimientos de inventario
        with db.connection:
            movimientos = pd.read_sql_query("""
                SELECT 
                    i.item,
                    i.categoria,
                    SUM(hc.cantidad) as total_vendido,
                    SUM(hc.precio_venta * hc.cantidad) as ingresos,
                    AVG(hc.margen) as margen_promedio
                FROM historial_consumo hc
                JOIN inventario i ON hc.id_repuesto = i.id
                WHERE hc.fecha >= DATE('now', '-30 days')
                GROUP BY i.id
                ORDER BY total_vendido DESC
                LIMIT 20
            """, db.connection)
        
        if not movimientos.empty:
            st.write("**Top 20 repuestos más vendidos (últimos 30 días)**")
            st.dataframe(movimientos, use_container_width=True)
    
    with tab3:
        st.subheader("Reportes de Clientes")
        
        with db.connection:
            top_clientes = pd.read_sql_query("""
                SELECT 
                    c.nombre,
                    c.telefono,
                    COUNT(r.id) as total_ordenes,
                    SUM(r.total) as total_gastado,
                    MAX(r.fecha_ingreso) as ultima_visita
                FROM clientes c
                LEFT JOIN reparaciones r ON c.id = r.cliente_id
                WHERE r.estado = 'Entregado'
                GROUP BY c.id
                ORDER BY total_gastado DESC
                LIMIT 20
            """, db.connection)
        
        if not top_clientes.empty:
            st.write("**Top 20 clientes por gasto**")
            st.dataframe(top_clientes, use_container_width=True)

def config_module():
    """Módulo de configuración del sistema"""
    st.markdown('<h1 class="main-header">⚙️ Configuración</h1>', unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["👥 Gestión de Usuarios", "🏢 Configuración General", "💾 Backups"])
    
    with tab1:
        st.subheader("Gestión de Usuarios")
        
        # Listar usuarios
        with db.connection:
            usuarios = pd.read_sql_query("""
                SELECT id, username, nombre_completo, email, rol, activo, ultimo_acceso
                FROM usuarios
                ORDER BY created_at DESC
            """, db.connection)
        
        st.dataframe(usuarios, use_container_width=True)
        
        # Crear nuevo usuario
        with st.expander("➕ Crear Nuevo Usuario"):
            with st.form("nuevo_usuario"):
                col1, col2 = st.columns(2)
                with col1:
                    new_username = st.text_input("Usuario *")
                    new_password = st.text_input("Contraseña *", type="password")
                    new_nombre = st.text_input("Nombre Completo *")
                with col2:
                    new_email = st.text_input("Email *")
                    new_rol = st.selectbox("Rol", ['admin', 'supervisor', 'tecnico', 'recepcion'])
                
                if st.form_submit_button("💾 Crear Usuario"):
                    if all([new_username, new_password, new_nombre, new_email]):
                        if auth_manager.create_user(new_username, new_password, new_nombre, new_email, new_rol):
                            st.success(f"✅ Usuario {new_username} creado correctamente")
                            st.rerun()
                        else:
                            st.error("❌ El usuario o email ya existe")
                    else:
                        st.error("Complete todos los campos obligatorios")
    
    with tab2:
        st.subheader("Configuración General")
        
        st.info("🔧 Configure las variables de entorno para personalizar el sistema:")
        
        config_items = {
            'SMTP_SERVER': 'Servidor de correo SMTP',
            'SMTP_USER': 'Usuario SMTP',
            'MP_ACCESS_TOKEN': 'Token de MercadoPago',
            'WHATSAPP_API_KEY': 'API Key de WhatsApp Business',
            'SESSION_TIMEOUT': 'Tiempo de sesión (minutos)'
        }
        
        for key, description in config_items.items():
            value = getattr(config, key, 'No configurado')
            masked = '✅ Configurado' if value and value not in ['', 'No configurado'] else '❌ No configurado'
            st.write(f"**{description}:** {masked}")
        
        st.caption("Estas configuraciones se establecen mediante variables de entorno o archivo .env")
    
    with tab3:
        st.subheader("Gestión de Backups")
        
        # Listar backups existentes
        backup_files = sorted([f for f in os.listdir(config.BACKUP_DIR) if f.endswith('.gz')], reverse=True)
        
        if backup_files:
            st.write(f"**Backups disponibles:** {len(backup_files)}")
            
            for backup in backup_files[:10]:  # Mostrar últimos 10
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(backup)
                with col2:
                    backup_path = os.path.join(config.BACKUP_DIR, backup)
                    with open(backup_path, 'rb') as f:
                        st.download_button(
                            "⬇️ Descargar",
                            f.read(),
                            file_name=backup,
                            key=f"dl_{backup}"
                        )
        else:
            st.info("No hay backups disponibles")
        
        if st.button("🔄 Forzar Backup Ahora", use_container_width=True):
            backup_file = db.backup_database(automatic=False)
            if backup_file:
                st.success(f"✅ Backup creado: {os.path.basename(backup_file)}")
                st.rerun()

def clientes_module():
    """Módulo de gestión de clientes"""
    st.markdown('<h1 class="main-header">👥 Gestión de Clientes</h1>', unsafe_allow_html=True)
    
    # Búsqueda de clientes
    busqueda_cliente = st.text_input("🔍 Buscar cliente por nombre, teléfono o DNI")
    
    query = """
        SELECT c.*, 
               COUNT(DISTINCT v.id) as total_vehiculos,
               COUNT(DISTINCT r.id) as total_ordenes,
               SUM(CASE WHEN r.estado = 'Entregado' THEN r.total ELSE 0 END) as total_gastado
        FROM clientes c
        LEFT JOIN vehiculos v ON c.id = v.cliente_id AND v.activo = 1
        LEFT JOIN reparaciones r ON c.id = r.cliente_id AND r.activo = 1
        WHERE c.activo = 1
    """
    params = []
    
    if busqueda_cliente:
        query += " AND (c.nombre LIKE ? OR c.telefono LIKE ? OR c.dni_cuit LIKE ?)"
        search_term = f"%{busqueda_cliente}%"
        params.extend([search_term, search_term, search_term])
    
    query += " GROUP BY c.id ORDER BY c.nombre"
    
    with db.connection:
        clientes = pd.read_sql_query(query, db.connection, params=params)
    
    if clientes.empty:
        st.info("No se encontraron clientes")
    else:
        st.dataframe(
            clientes,
            column_config={
                "total_gastado": st.column_config.NumberColumn("Total Gastado", format="$%.2f")
            },
            use_container_width=True,
            height=400
        )
        
        # Ver detalle de cliente
        cliente_seleccionado = st.selectbox("Seleccionar cliente para ver detalle", 
                                          clientes['nombre'].tolist() if not clientes.empty else [])
        
        if cliente_seleccionado:
            cliente_id = clientes[clientes['nombre'] == cliente_seleccionado].iloc[0]['id']
            
            with db.connection:
                # Vehículos del cliente
                vehiculos = pd.read_sql_query("""
                    SELECT * FROM vehiculos WHERE cliente_id = ? AND activo = 1
                """, db.connection, params=(cliente_id,))
                
                # Historial de órdenes
                ordenes = pd.read_sql_query("""
                    SELECT r.*, v.patente 
                    FROM reparaciones r
                    LEFT JOIN vehiculos v ON r.vehiculo_id = v.id
                    WHERE r.cliente_id = ? AND r.activo = 1
                    ORDER BY r.fecha_ingreso DESC
                """, db.connection, params=(cliente_id,))
            
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("🚗 Vehículos")
                if not vehiculos.empty:
                    st.dataframe(vehiculos[['patente', 'marca', 'modelo', 'anio']], use_container_width=True)
                else:
                    st.info("Sin vehículos registrados")
            
            with col2:
                st.subheader("📋 Historial de Servicios")
                if not ordenes.empty:
                    st.dataframe(ordenes[['numero_orden', 'fecha_ingreso', 'estado', 'total']], use_container_width=True)
                else:
                    st.info("Sin órdenes registradas")

# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    """Punto de entrada principal"""
    
    # Verificar autenticación
    if "authenticated" not in st.session_state or not st.session_state["authenticated"]:
        login_screen()
        return
    
    # Menú lateral y routing
    selected_module = sidebar_menu()
    
    # Routing de módulos
    if selected_module == "dashboard":
        dashboard_module()
    elif selected_module == "inventario":
        inventario_module()
    elif selected_module == "clientes":
        clientes_module()
    elif selected_module == "nueva_orden":
        nueva_orden_module()
    elif selected_module == "taller":
        taller_module()
    elif selected_module == "facturacion":
        facturacion_module()
    elif selected_module == "reportes":
        reportes_module()
    elif selected_module == "config":
        config_module()

if __name__ == "__main__":
    main()
            
